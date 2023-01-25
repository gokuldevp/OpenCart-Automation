import logging
import configparser
import datetime
import os
import openpyxl
from openpyxl.styles import PatternFill


def get_current_date_time():
    """function to return datetime int YYMMDDHHmmSS format"""
    current_datetime = str(datetime.datetime.now())[0:19]
    return datetime.datetime.strptime(current_datetime, '%Y-%m-%d %H:%M:%S').strftime('%Y%m%d%H%M%S')


class CreateUser:
    """Class to create the test user information."""

    @staticmethod
    def create_email():
        """Method to return email address based on the current timestamp"""
        email = f"ttpvt_gd{get_current_date_time()}@yopmail.com"
        return email


class ScreenShots:
    """Class to take screenshots"""

    @staticmethod
    def take_screenshots_as_png(driver, screenshot_name):
        """
        Method to take screenshot, Provide the driver, screenshot name
        This method will return the screenshot in png format
        """
        screenshot_name += " " + datetime.datetime.strptime(get_current_date_time(), '%Y%m%d%H%M%S').strftime(
            '%d_%m_%Y %H-%M-%S') + ".png"
        screenshot_file = os.path.join(os.path.dirname(os.path.abspath(__file__).replace('utilities\\', '')),
                                       'screenshots', screenshot_name)
        driver.save_screenshot(screenshot_file)


config = configparser.RawConfigParser()
config.read(
    os.path.join(os.path.dirname(os.path.abspath(__file__).replace('utilities\\', '')), 'configurations', 'config.ini'))


class ReadConfig:
    """This is a class to get the details in the config.ini file."""

    @staticmethod
    def get_application_url():
        return config.get("commonInfo", "baseURL")

    @staticmethod
    def get_username():
        return config.get("commonInfo", "email")

    @staticmethod
    def get_password():
        return config.get("commonInfo", "password")


class LogGen:
    """Class to Generate Logs"""

    # Note: The level priority are as follows Warn> Debug> Info> Error> Fatal
    @staticmethod
    def loggen():
        log_file = os.path.join(os.path.dirname(os.path.abspath(__file__).replace('utilities\\', '')), 'logs',
                                'automation.log')
        log_handler = logging.FileHandler(log_file)
        log_handler.setFormatter(
            logging.Formatter("%(asctime)s: %(levelname)s: %(message)s", datefmt="%m/%d/%Y %I:%M:%S %p"))
        logger = logging.getLogger()
        logger.addHandler(log_handler)
        logger.setLevel(logging.INFO)
        return logger


class Excel:
    """Class to handle Excel file"""

    @staticmethod
    def get_row_count(file, sheet_name):
        """Method to return the row count of the Excel file"""
        workbook = openpyxl.load_workbook(file)
        sheet = workbook[sheet_name]
        return sheet.max_row

    @staticmethod
    def get_column_count(file, sheet_name):
        """Method to return the column count of the Excel file"""
        workbook = openpyxl.load_workbook(file)
        sheet = workbook[sheet_name]
        return sheet.max_column

    @staticmethod
    def read_data(file, sheet_name, row_num, column_num):
        """Method to read data from Excel file"""
        workbook = openpyxl.load_workbook(file)
        sheet = workbook[sheet_name]
        return sheet.cell(row_num, column_num).value

    @staticmethod
    def write_data(file, sheet_name, row_num, column_num, data):
        """Method to write data from Excel file"""
        workbook = openpyxl.load_workbook(file)
        sheet = workbook[sheet_name]
        sheet.cell(row_num, column_num).value = data
        workbook.save(file)

    @staticmethod
    def fill_green_color(file, sheet_name, row_num, column_num):
        """Method to add green color to Excel cell text"""
        workbook = openpyxl.load_workbook(file)
        sheet = workbook[sheet_name]
        green_fill = PatternFill(start_color='60b212',
                                 end_color='60b212',
                                 fill_type='solid')
        sheet.cell(row_num, column_num).fill = green_fill
        workbook.save(file)

    @staticmethod
    def fill_red_color(file, sheet_name, row_num, column_num):
        """Method to add red color to Excel cell text"""
        workbook = openpyxl.load_workbook(file)
        sheet = workbook[sheet_name]
        red_fill = PatternFill(start_color='ff0000',
                               end_color='ff0000',
                               fill_type='solid')
        sheet.cell(row_num, column_num).fill = red_fill
        workbook.save(file)


class TestData:
    """Class to get test data file"""

    @staticmethod
    def get_login_test_data():
        """Method to return the login Excel file"""
        data_file = os.path.join(os.path.dirname(os.path.abspath(__file__).replace('utilities\\', '')), 'testdata',
                                 'Opencart_LoginData.xlsx')
        return data_file
